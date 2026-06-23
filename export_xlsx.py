"""
export_xlsx.py - Rich XLSX Export for Ranked Candidates
=======================================================
Redrob Data & AI Challenge - Export Module

Exports the top-100 ranked candidates to a professionally formatted .xlsx
workbook with two sheets:

  Sheet 1 — "Top 100 Candidates":
      Header row with bold white text on dark-blue background, auto-filter,
      conditional 3-color scales on score columns, gold/silver rank highlights,
      alternating row colors, frozen header, and auto-adjusted column widths.

  Sheet 2 — "Summary":
      Generation timestamp, total candidates, score statistics (min / max /
      mean), and optional team metadata.

Usage:
    from export_xlsx import export_to_xlsx

    export_to_xlsx(rows, "output.xlsx", metadata={"team_name": "Redrob"})

Dependencies:
    openpyxl (optional — graceful fallback if not installed)
"""

import os
from datetime import datetime


# ──────────────────────────────────────────────
# OPTIONAL DEPENDENCY GUARD
# ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ──────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────
SHEET_NAME       = "Top 100 Candidates"
SUMMARY_SHEET    = "Summary"

# Colours (ARGB hex without leading #)
HEADER_BG        = "1B2A4A"
HEADER_FG        = "FFFFFF"
GOLD_BG          = "FFD700"
SILVER_BG        = "C0C0C0"
ALT_ROW_GRAY     = "F2F2F2"
ALT_ROW_WHITE    = "FFFFFF"

# Score color-scale endpoints (red → yellow → green)
SCALE_RED        = "F8696B"
SCALE_YELLOW     = "FFEB84"
SCALE_GREEN      = "63BE7B"

# Column definitions: (key, header_label, width)
# Optional columns are appended dynamically if present in the data.
_BASE_COLUMNS = [
    ("rank",       "Rank",       8),
    ("candidate_id", "Candidate ID", 20),
    ("score",      "Score",     12),
]

_OPTIONAL_COLUMNS = [
    ("semantic_score", "Semantic Score", 16),
    ("skill_score",    "Skill Score",    14),
    ("career_score",   "Career Score",   14),
    ("multiplier",     "Multiplier",     12),
]

_TAIL_COLUMNS = [
    ("reasoning", "Reasoning", 80),
]

# Number format for score cells
SCORE_FMT = "0.0000"


# ──────────────────────────────────────────────
# STYLE HELPERS
# ──────────────────────────────────────────────
def _header_font():
    """Bold white header font."""
    return Font(name="Calibri", bold=True, color=HEADER_FG, size=11)


def _header_fill():
    """Dark-blue header background."""
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG,
                       fill_type="solid")


def _rank_fill(rank: int):
    """Return gold fill for top-3, silver for 4-10, else None."""
    if rank <= 3:
        return PatternFill(start_color=GOLD_BG, end_color=GOLD_BG,
                           fill_type="solid")
    if rank <= 10:
        return PatternFill(start_color=SILVER_BG, end_color=SILVER_BG,
                           fill_type="solid")
    return None


def _alt_fill(row_idx: int):
    """Alternating row fill — gray for even data rows, white for odd."""
    color = ALT_ROW_GRAY if row_idx % 2 == 0 else ALT_ROW_WHITE
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _thin_border():
    """Light thin border for all cells."""
    side = Side(style="thin", color="D9D9D9")
    return Border(top=side, bottom=side, left=side, right=side)


# ──────────────────────────────────────────────
# COLUMN RESOLUTION
# ──────────────────────────────────────────────
def _resolve_columns(rows):
    """
    Build the ordered column list based on which optional keys are
    actually present in the first data row.

    Returns
    -------
    list of (key, label, width) tuples
    """
    sample = rows[0] if rows else {}
    columns = list(_BASE_COLUMNS)
    for col in _OPTIONAL_COLUMNS:
        if col[0] in sample:
            columns.append(col)
    columns.extend(_TAIL_COLUMNS)
    return columns


# ──────────────────────────────────────────────
# SCORE STATISTICS
# ──────────────────────────────────────────────
def _score_stats(rows):
    """
    Compute min, max, and mean of the 'score' field.

    Returns
    -------
    dict  with keys min, max, mean  (or empty dict if no scores).
    """
    scores = [r["score"] for r in rows if "score" in r and r["score"] is not None]
    if not scores:
        return {}
    return {
        "min":  min(scores),
        "max":  max(scores),
        "mean": sum(scores) / len(scores),
    }


# ──────────────────────────────────────────────
# MAIN EXPORT FUNCTION
# ──────────────────────────────────────────────
def export_to_xlsx(rows: list, output_path: str,
                   metadata: dict = None) -> str:
    """
    Export ranked candidates to a richly formatted XLSX workbook.

    Parameters
    ----------
    rows : list[dict]
        Each dict must contain at least ``candidate_id``, ``rank``,
        ``score``, and ``reasoning``.  Optional keys: ``semantic_score``,
        ``skill_score``, ``career_score``, ``multiplier``.
    output_path : str
        Destination path for the .xlsx file.
    metadata : dict, optional
        Extra info written to the Summary sheet (e.g. ``team_name``,
        ``timestamp``).

    Returns
    -------
    str   absolute path of the written file, or None on failure.
    """
    # ── graceful fallback ────────────────────
    if not _HAS_OPENPYXL:
        print("[WARNING] openpyxl is not installed — XLSX export skipped. "
              "Install it with:  pip install openpyxl")
        return None

    if not rows:
        print("[WARNING] No rows to export — XLSX export skipped.")
        return None

    metadata = metadata or {}
    columns  = _resolve_columns(rows)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Top 100 Candidates ──────────
    ws = wb.active
    ws.title = SHEET_NAME

    # Identify which columns hold numeric scores (for conditional formatting)
    score_keys = {"score", "semantic_score", "skill_score",
                  "career_score", "multiplier"}
    score_col_indices = []          # 1-based column indices

    # ── header row ───────────────────────────
    for col_idx, (key, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = _header_font()
        cell.fill      = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        if key in score_keys:
            score_col_indices.append(col_idx)

    # ── data rows ────────────────────────────
    for row_offset, record in enumerate(rows, start=2):
        rank_val  = record.get("rank")
        rank_fill = _rank_fill(rank_val) if isinstance(rank_val, (int, float)) else None
        alt_fill  = _alt_fill(row_offset)

        for col_idx, (key, _label, _w) in enumerate(columns, start=1):
            value = record.get(key, "")
            cell  = ws.cell(row=row_offset, column=col_idx, value=value)

            # Number formatting for score columns
            if key in score_keys and isinstance(value, (int, float)):
                cell.number_format = SCORE_FMT

            # Alignment
            if key == "reasoning":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")

            # Fill: rank highlight trumps alternating color
            if key == "rank" and rank_fill is not None:
                cell.fill = rank_fill
            else:
                cell.fill = alt_fill

            cell.border = _thin_border()

    # ── conditional formatting (3-color scale) on score columns ──
    last_data_row = len(rows) + 1
    for ci in score_col_indices:
        col_letter = get_column_letter(ci)
        cell_range = f"{col_letter}2:{col_letter}{last_data_row}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="min", start_color=SCALE_RED,
                mid_type="percentile", mid_value=50, mid_color=SCALE_YELLOW,
                end_type="max", end_color=SCALE_GREEN,
            ),
        )

    # ── auto-filter & freeze panes ───────────
    last_col_letter = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"
    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ─────────────────────
    _write_summary_sheet(wb, rows, metadata)

    # ── save workbook ────────────────────────
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    abs_path = os.path.abspath(output_path)
    print(f"[INFO] XLSX exported -> {abs_path}")
    return abs_path


# ──────────────────────────────────────────────
# SUMMARY SHEET
# ──────────────────────────────────────────────
def _write_summary_sheet(wb, rows, metadata):
    """
    Create the 'Summary' sheet with generation stats and metadata.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Target workbook (sheet is appended).
    rows : list[dict]
        Ranked candidate rows.
    metadata : dict
        Optional metadata dict.
    """
    ws = wb.create_sheet(title=SUMMARY_SHEET)

    # Styling helpers
    title_font  = Font(name="Calibri", bold=True, size=14, color=HEADER_BG)
    label_font  = Font(name="Calibri", bold=True, size=11)
    value_font  = Font(name="Calibri", size=11)
    label_fill  = PatternFill(start_color="E8EDF3", end_color="E8EDF3",
                              fill_type="solid")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    # ── title ────────────────────────────────
    ws.cell(row=1, column=1, value="Ranking Summary").font = title_font

    # ── summary rows ─────────────────────────
    stats = _score_stats(rows)
    timestamp = metadata.get("timestamp", datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"))

    summary_items = [
        ("Generation Timestamp", timestamp),
        ("Total Candidates Ranked", len(rows)),
    ]

    if stats:
        summary_items.extend([
            ("Score — Min",  f"{stats['min']:.4f}"),
            ("Score — Max",  f"{stats['max']:.4f}"),
            ("Score — Mean", f"{stats['mean']:.4f}"),
        ])

    team_name = metadata.get("team_name")
    if team_name:
        summary_items.append(("Team Name", team_name))

    # Write any remaining metadata keys not already covered
    _covered = {"timestamp", "team_name"}
    for k, v in metadata.items():
        if k not in _covered:
            summary_items.append((str(k), str(v)))

    for idx, (label, value) in enumerate(summary_items, start=3):
        lc = ws.cell(row=idx, column=1, value=label)
        lc.font = label_font
        lc.fill = label_fill
        lc.border = _thin_border()

        vc = ws.cell(row=idx, column=2, value=value)
        vc.font   = value_font
        vc.border = _thin_border()


# ──────────────────────────────────────────────
# CLI CONVENIENCE
# ──────────────────────────────────────────────
if __name__ == "__main__":
    # Quick smoke-test with dummy data
    sample_rows = [
        {
            "rank": i,
            "candidate_id": f"CAND-{i:05d}",
            "score": round(1.0 - i * 0.008, 4),
            "semantic_score": round(0.9 - i * 0.005, 4),
            "skill_score": round(0.85 - i * 0.006, 4),
            "career_score": round(0.8 - i * 0.004, 4),
            "multiplier": round(1.0 + (0.02 if i <= 5 else 0.0), 4),
            "reasoning": f"Strong match on core skills. Rank {i} candidate.",
        }
        for i in range(1, 101)
    ]

    out = export_to_xlsx(
        sample_rows,
        os.path.join(os.path.dirname(__file__), "ranked_candidates.xlsx"),
        metadata={"team_name": "Redrob", "challenge": "India Runs 2026"},
    )
    if out:
        print(f"Smoke-test passed — {out}")
